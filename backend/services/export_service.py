"""CSV/Excel/PDF export helpers for Portfolio Dashboard and Accounting reports.
Shared, format-agnostic table export (CSV/XLSX) plus two purpose-built PDF
builders that reuse the same reportlab styling already established in
document_generator.py.
"""
import csv
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

ACCENT = colors.HexColor('#B8860B')
HEADER_BG = colors.HexColor('#1E293B')


def rows_to_csv(rows: list, headers: list = None) -> bytes:
    if not rows:
        return b""
    headers = headers or list(rows[0].keys())
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode('utf-8')


_EXCEL_PRIMITIVES = (str, int, float, bool, type(None))


def _excel_safe(value):
    return value if isinstance(value, _EXCEL_PRIMITIVES) else str(value)


def sheets_to_excel(sheets: dict) -> bytes:
    """sheets: {sheet_name: [row_dict, ...]}"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([_excel_safe(row.get(h)) for h in headers])
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _base_doc(buf, title: str):
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ExportTitle', parent=styles['Title'], textColor=ACCENT, fontSize=18)
    meta_style = ParagraphStyle('ExportMeta', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#64748B'))
    elements = [
        Paragraph(title, title_style),
        Paragraph(f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", meta_style),
        Spacer(1, 16),
    ]
    return doc, elements, styles


def _table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])


def portfolio_dashboard_pdf(property_rows: list, totals: dict, org_name: str) -> bytes:
    buf = io.BytesIO()
    doc, elements, styles = _base_doc(buf, f"Portfolio Dashboard — {org_name}")

    summary_headers = ["Properties", "Units", "Occupied", "Revenue MRR", "Collected 30d", "Delinquent $", "Maint. Open"]
    summary_row = [
        totals.get("total_properties"), totals.get("total_units"), totals.get("occupied_units"),
        f"${totals.get('revenue_mrr', 0):,.0f}", f"${totals.get('collected_30d', 0):,.0f}",
        f"${totals.get('delinquent_amount', 0):,.0f}", totals.get("maintenance_open"),
    ]
    summary_table = Table([summary_headers, summary_row], hAlign='LEFT')
    summary_table.setStyle(_table_style())
    elements += [summary_table, Spacer(1, 20)]

    prop_headers = ["Property", "Units", "Occ.", "Vacancy %", "Rev MRR", "Cash Flow 30d", "Delinquent", "Health"]
    prop_rows = [prop_headers]
    for p in property_rows:
        prop_rows.append([
            p["property_name"], p["total_units"], p["occupied_units"], f"{p['vacancy_rate']}%",
            f"${p['revenue_mrr']:,.0f}", f"${p['cash_flow_30d']:,.0f}",
            f"${p['delinquent_amount']:,.0f}", f"{p['health_score']} ({p['health_grade']})",
        ])
    prop_table = Table(prop_rows, hAlign='LEFT', repeatRows=1)
    prop_table.setStyle(_table_style())
    elements.append(prop_table)

    doc.build(elements)
    return buf.getvalue()


def accounting_report_pdf(report: dict, scope_label: str) -> bytes:
    buf = io.BytesIO()
    doc, elements, styles = _base_doc(buf, f"Financial Report — {scope_label}")

    period_style = ParagraphStyle('Period', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#334155'))
    elements.append(Paragraph(f"Period: {report.get('period_start')} to {report.get('period_end')}", period_style))
    elements.append(Spacer(1, 14))

    rows = [
        ["Metric", "Value"],
        ["Total Income", f"${report.get('total_income', 0):,.2f}"],
        ["Total Expenses", f"${report.get('total_expenses', 0):,.2f}"],
        ["Operating Expenses", f"${report.get('operating_expenses', 0):,.2f}"],
        ["Net Operating Income (NOI)", f"${report.get('noi', 0):,.2f}"],
        ["Annualized NOI", f"${report.get('annualized_noi', 0):,.2f}"],
        ["Cap Rate", f"{report.get('cap_rate')}%" if report.get('cap_rate') is not None else "N/A"],
        ["DSCR", str(report.get('dscr')) if report.get('dscr') is not None else "N/A"],
        ["Cash Flow", f"${report.get('cash_flow', 0):,.2f}"],
    ]
    table = Table(rows, hAlign='LEFT', colWidths=[220, 150])
    table.setStyle(_table_style())
    elements.append(table)

    doc.build(elements)
    return buf.getvalue()
